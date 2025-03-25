import requests
from lxml import etree
from openpyxl import Workbook
from tqdm import tqdm
from openpyxl import load_workbook
import concurrent.futures


def fetch_color_details(url):
    """抓取颜色页面的详细信息（修复异常抛出逻辑）"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    }

    try:
        with requests.Session() as session:
            response = session.get(url, headers=headers, timeout=(5, 15))
            response.raise_for_status()  # 自动处理HTTP错误码

            tree = etree.HTML(response.content)

            def get_value(label):
                xpath = f'//tr[td[@class="left" and contains(., "{label}")]]/td[@class="right"]'
                if elements := tree.xpath(xpath):
                    return elements[0].xpath('string()').strip()
                raise etree.XPathError(f"未找到 {label} 对应元素")

            # 关键数据校验
            hex_code = get_value("Hex Code")
            if not hex_code.startswith("#"):
                raise ValueError(f"无效的Hex格式: {hex_code}")

            return {
                'Hex Code': hex_code,
                'RGB Values': get_value("RGB Values"),
                'CMYK Values': get_value("CMYK Values"),
                'HSV/HSB Values': get_value("HSV/HSB Values"),
                'Closest RAL': get_value("Closest RAL")
            }

    except Exception as e:
        # 将原始异常重新抛出，保留堆栈信息
        raise type(e)(f"{url} 处理失败: {str(e)}") from e


def save_failed_urls(failed_urls):
    """保存失败记录到单独的工作表"""
    if not failed_urls:
        return

    try:
        # 尝试追加到现有文件
        wb = load_workbook('color_details.xlsx')
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认sheet

    # 创建失败记录工作表
    ws = wb.create_sheet("失败记录")
    ws.append(['URL', '错误类型', '错误详情'])

    # 分类统计错误类型
    error_stats = {}
    for (url, error_type, error_detail) in failed_urls:
        error_stats[error_type] = error_stats.get(error_type, 0) + 1
        ws.append([url, error_type, error_detail])

    # 添加统计信息
    ws.append([])
    ws.append(['错误类型', '出现次数'])
    for error_type, count in error_stats.items():
        ws.append([error_type, count])

    wb.save('color_details.xlsx')
    print(f"已保存{len(failed_urls)}条失败记录到color_details.xlsx的'失败记录'工作表")


def save_to_excel(all_data):
    """保存到Excel文件（增加数据校验）"""
    # 过滤非字典类型的数据
    valid_data = {url: details for url, details in all_data.items() if isinstance(details, dict)}

    wb = Workbook()
    ws = wb.active
    ws.title = "颜色代码"

    headers = ['序号', '颜色链接', 'Hex Code', 'RGB Values', 'CMYK Values',
               'HSV/HSB Values', 'Closest RAL']
    ws.append(headers)

    # 设置列宽优化显示
    column_widths = {
        'B': 90,  # 链接列
        'C': 20,  # Hex列
        'D': 30,  # RGB列
        'E': 30,  # CMYK列
        'F': 30,  # HSV列
        'G': 30  # RAL列
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 填充数据
    for idx, (url, details) in enumerate(valid_data.items(), start=1):
        # 添加空值保护
        row = [
            idx,
            url,
            details.get('Hex Code', 'N/A'),
            details.get('RGB Values', 'N/A'),
            details.get('CMYK Values', 'N/A'),
            details.get('HSV/HSB Values', 'N/A'),
            details.get('Closest RAL', 'N/A')
        ]
        ws.append(row)

    wb.save('color_details.xlsx')
    print(f"已保存{len(valid_data)}条有效数据到color_details.xlsx")


def load_links_from_excel(file_path='color_links.xlsx'):
    """从Excel读取链接列表"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        return [row[1].value for row in ws.iter_rows(min_row=2) if row[1].value]
    except Exception as e:
        print(f"读取链接失败: {str(e)}")
        return []


if __name__ == '__main__':
    color_links = load_links_from_excel()
    if not color_links:
        print("未找到有效链接")
        exit()

    all_details = {}
    failed_urls = []

    # 使用线程池（限制最大并发数为10）
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        future_to_url = {executor.submit(fetch_color_details, url): url for url in color_links}

        # 进度条设置（优化显示单位）
        with tqdm(concurrent.futures.as_completed(future_to_url),
                  total=len(future_to_url),
                  desc="抓取进度",
                  unit="个") as pbar:

            for future in pbar:
                url = future_to_url[future]
                try:
                    result = future.result()
                    # 添加结果类型校验
                    if isinstance(result, dict):
                        all_details[url] = result
                    else:
                        raise ValueError("返回非字典类型结果")
                except Exception as e:
                    error_type = type(e).__name__
                    error_msg = str(e).split(": ")[-1]  # 去除URL前缀
                    failed_urls.append((url, error_type, error_msg))

        # 保存成功数据
    if all_details:
        save_to_excel(all_details)

        # 保存失败记录
    if failed_urls:
        save_failed_urls(failed_urls)

        # 打印最终统计
    print(f"\n最终统计：")
    print(f"成功抓取: {len(all_details)}条")
    print(f"失败记录: {len(failed_urls)}条")
    if failed_urls:
        print("失败原因分类：")
        from collections import defaultdict

        stats = defaultdict(int)
        for _, error_type, _ in failed_urls:
            stats[error_type] += 1
        for error_type, count in stats.items():
            print(f"  {error_type}: {count}次")
