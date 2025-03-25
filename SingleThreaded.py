import requests
from lxml import etree
from openpyxl import Workbook
from tqdm import tqdm  # 进度条支持
from openpyxl import load_workbook

def fetch_color_details(url):
    """抓取颜色页面的详细信息"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    }

    try:
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code != 200:
            return None

        tree = etree.HTML(response.content)

        # 使用更健壮的XPath定位方式
        def get_value(label):
            xpath = f'//tr[td[@class="left" and contains(., "{label}")]]/td[@class="right"]'
            element = tree.xpath(xpath)
            return element[0].xpath('string()').strip() if element else "N/A"

        return {
            'Hex Code': get_value("Hex Code"),
            'RGB Values': get_value("RGB Values"),  # 保留原始空格
            'CMYK Values': get_value("CMYK Values"),  # 保留百分号
            'HSV/HSB Values': get_value("HSV/HSB Values"),  # 保留角度符号°
            'Closest RAL': get_value("Closest RAL")  # 保留完整描述
        }

    except Exception as e:
        print(f"抓取失败：{url}，错误：{str(e)}")
        return None


def save_to_excel(all_data):
    """保存到Excel文件（合并后的版本）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "颜色代码"

    # 创建带格式的表头
    headers = ['序号', '颜色链接', 'Hex Code', 'RGB Values', 'CMYK Values',
               'HSV/HSB Values', 'Closest RAL']
    ws.append(headers)

    # 设置列宽（优化显示效果）
    ws.column_dimensions['B'].width = 50  # 链接列
    ws.column_dimensions['C'].width = 12  # Hex列
    ws.column_dimensions['D'].width = 20  # RGB列

    # 填充数据并添加错误处理
    for idx, (url, details) in enumerate(all_data.items(), start=1):
        try:
            row = [
                idx,
                url,
                details.get('Hex Code', ''),
                details.get('RGB Values', ''),  # 原始格式如：(164, 119, 100)
                details.get('CMYK Values', ''),  # 原始格式如：(0%, 27%, 39%, 36%)
                details.get('HSV/HSB Values', ''),  # 原始格式如：18°, 39%, 64%
                details.get('Closest RAL', '')  # 原始格式如：1011 [Brown beige]
            ]
            ws.append(row)
        except Exception as e:
            print(f"数据写入异常：{url}，错误：{str(e)}")

    wb.save('color_details.xlsx')
    print(f"已保存{len(all_data)}条数据到color_details.xlsx")

def save_to_excel(all_data):
    """保存到Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "颜色代码"

    # 创建表头
    headers = ['序号', '颜色链接', 'Hex Code', 'RGB Values', 'CMYK Values', 'HSV/HSB Values', 'Closest RAL']
    ws.append(headers)

    # 填充数据
    for idx, (url, details) in enumerate(all_data.items(), start=1):
        row = [idx, url] + [details.get(h, '') for h in headers[2:]]
        ws.append(row)

    wb.save('color_details.xlsx')
    print(f"已保存{len(all_data)}条数据到color_details.xlsx")


def load_links_from_excel(file_path='color_links.xlsx'):
    """从之前生成的Excel文件中读取所有颜色链接"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        # 获取B列（颜色链接列）的所有值，跳过表头
        links = [row[1].value for row in ws.iter_rows(min_row=2) if row[1].value]
        return links
    except FileNotFoundError:
        print(f"错误：找不到文件 {file_path}")
        return []
    except Exception as e:
        print(f"读取Excel异常：{str(e)}")
        return []


if __name__ == '__main__':
    # 从Excel加载所有链接
    color_links = load_links_from_excel()

    if not color_links:
        print("未找到有效链接，请确保已生成color_links.xlsx")
        exit()

    all_details = {}
    for link in tqdm(color_links, desc="抓取进度"):
        details = fetch_color_details(link)
        if details:
            all_details[link] = details

    if all_details:
        save_to_excel(all_details)
    else:
        print("未获取到有效数据")