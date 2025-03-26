import requests
from requests.adapters import HTTPAdapter
from lxml import etree
from openpyxl import Workbook
from fake_useragent import UserAgent
from openpyxl import load_workbook

# 常量配置
BASE_URL = 'https://www.color-name.com/search/{color}'
EXCEL_PATH = 'colorRal_links.xlsx'
EXCEL_COLORS_PATH = 'colorral.xlsx'
HEADERS = {'User-Agent': UserAgent().random}


def load_colors_from_excel(file_path: str) -> list[str]:
    """从 Excel 文件中提取颜色名称列表"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        colors = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            color = row[0]
            if color and isinstance(color, str):
                colors.append(color.strip().lower())
        return colors
    except Exception as e:
        print(f"读取颜色文件失败: {e}")
        return []


COLORS = load_colors_from_excel(EXCEL_COLORS_PATH)


def get_color_links(color: str, max_retries: int = 3) -> list[str]:
    """根据颜色名称生成动态URL并抓取链接"""
    session = requests.Session()
    session.mount('https://', HTTPAdapter(max_retries=max_retries))

    try:
        url = BASE_URL.format(color=color.lower())
        response = session.get(url, headers=HEADERS, timeout=10)
        response.raise_for_status()
        tree = etree.HTML(response.content)

        # 修改后的XPath：提取href属性
        links = tree.xpath('/html/body/div[2]/ul/li[1]/a/@href')  # 添加/@href
        return links if links else []
    except Exception as e:
        print(f"Error fetching {color}: {e}")
        return []


def save_to_excel(all_links: list[str]) -> None:
    """去重后保存到Excel"""
    unique_links = list(dict.fromkeys(all_links))  # 去重保留顺序

    wb = Workbook()
    ws = wb.active
    ws.title = "颜色链接"
    ws.append(['序号', '颜色链接'])

    for idx, link in enumerate(unique_links, start=1):
        ws.append([idx, link])

    wb.save(EXCEL_PATH)
    wb.close()


if __name__ == '__main__':
    if not COLORS:
        print("颜色列表为空，请检查 Excel 文件！")
        exit()

    all_links = []

    for color in COLORS:
        print(f"正在抓取 {color}...")
        links = get_color_links(color)
        if links:
            all_links.extend(links)

    if all_links:
        save_to_excel(all_links)
        print(f"已保存{len(all_links)}条数据到{EXCEL_PATH}（去重后{len(set(all_links))}条）")
    else:
        print("未获取到有效链接")