import requests
from lxml import etree
from openpyxl import Workbook
from tqdm import tqdm
from openpyxl import load_workbook
import concurrent.futures


def fetch_color_details(url):
    """抓取颜色页面的详细信息（使用Session优化）"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    }

    try:
        # 使用Session保持连接
        with requests.Session() as session:
            response = session.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                return None

            tree = etree.HTML(response.content)

            def get_value(label):
                xpath = f'//tr[td[@class="left" and contains(., "{label}")]]/td[@class="right"]'
                element = tree.xpath(xpath)
                return element[0].xpath('string()').strip() if element else "N/A"

            return {
                'Hex Code': get_value("Hex Code"),
                'RGB Values': get_value("RGB Values"),
                'CMYK Values': get_value("CMYK Values"),
                'HSV/HSB Values': get_value("HSV/HSB Values"),
                'Closest RAL': get_value("Closest RAL")
            }

    except Exception as e:
        # print(f"抓取失败：{url}，错误：{str(e)}")  # 避免过多错误输出
        return None


def save_to_excel(all_data):
    """保存到Excel文件（优化列宽）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "颜色代码"

    # 创建带格式的表头
    headers = ['序号', '颜色链接', 'Hex Code', 'RGB Values', 'CMYK Values',
               'HSV/HSB Values', 'Closest RAL']
    ws.append(headers)

    # 设置列宽优化显示
    column_widths = {
        'B': 50,  # 链接列
        'C': 12,  # Hex列
        'D': 20,  # RGB列
        'E': 20,  # CMYK列
        'F': 20,  # HSV列
        'G': 20  # RAL列
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 填充数据
    for idx, (url, details) in enumerate(all_data.items(), start=1):
        row = [
            idx,
            url,
            details.get('Hex Code', ''),
            details.get('RGB Values', ''),
            details.get('CMYK Values', ''),
            details.get('HSV/HSB Values', ''),
            details.get('Closest RAL', '')
        ]
        ws.append(row)

    wb.save('color_details.xlsx')
    print(f"已保存{len(all_data)}条数据到color_details.xlsx")


def load_links_from_excel(file_path='color_links.xlsx'):
    """从Excel读取链接列表"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        return [row[1].value for row in ws.iter_rows(min_row=2) if row[1].value]
    except Exception as e:
        print(f"读取链接失败: {str(e)}")
        return []


if __name__ == '__main__':
    color_links = load_links_from_excel()
    if not color_links:
        print("未找到有效链接")
        exit()

    all_details = {}

    # 使用线程池并发处理（默认workers数为10）
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = {executor.submit(fetch_color_details, url): url for url in color_links}

        # 使用tqdm显示进度条
        for future in tqdm(concurrent.futures.as_completed(futures),
                           total=len(futures),
                           desc="抓取进度",
                           unit="page"):
            url = futures[future]
            try:
                if (result := future.result()):
                    all_details[url] = result
            except Exception as e:
                pass  # 错误信息已在fetch函数中处理

    if all_details:
        save_to_excel(all_details)
    else:
        print("未获取到有效数据")